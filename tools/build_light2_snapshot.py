#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import math
import warnings
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def format_number(value: float) -> str:
    if math.isfinite(value) and float(value).is_integer():
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def infer_kind(value: Any, number_format: str) -> str:
    fmt = (number_format or "").lower()
    if isinstance(value, (datetime, date, time)):
        return "date"
    if isinstance(value, (int, float)):
        if "%" in fmt:
            return "percent"
        return "number"
    return "text"


def serialize_raw(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            if float(value).is_integer():
                return int(value)
            return round(float(value), 6)
        return None
    if value is None:
        return ""
    return str(value)


def format_display(value: Any, number_format: str) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")

    if isinstance(value, bool):
        return "Да" if value else "Нет"

    if isinstance(value, (int, float)):
        fmt = (number_format or "").lower()
        numeric = float(value)
        if "%" in fmt:
          return f"{numeric * 100:.2f}%".replace(".", ",")
        return format_number(numeric).replace(".", ",")

    return str(value)


def build_snapshot(source_path: Path) -> dict[str, Any]:
    workbook_formula = load_workbook(source_path, data_only=False)
    workbook_values = load_workbook(source_path, data_only=True)

    sheets_payload: list[dict[str, Any]] = []

    for sheet_formula in workbook_formula.worksheets:
        sheet_values = workbook_values[sheet_formula.title]

        sheet_rows: list[dict[str, Any]] = []
        non_empty = 0
        formulas = 0
        max_col = 1

        for row_index in range(1, sheet_formula.max_row + 1):
            row_cells: dict[str, Any] = {}

            for column_index in range(1, sheet_formula.max_column + 1):
                formula_cell = sheet_formula.cell(row=row_index, column=column_index)
                value_cell = sheet_values.cell(row=row_index, column=column_index)

                formula = ""
                if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                    formula = formula_cell.value

                raw_value = value_cell.value
                display_value = format_display(raw_value, value_cell.number_format or formula_cell.number_format)
                raw_serialized = serialize_raw(raw_value)

                if raw_serialized == "" and display_value == "" and not formula:
                    continue

                non_empty += 1
                if formula:
                    formulas += 1

                max_col = max(max_col, column_index)
                payload = {
                    "display": display_value,
                    "raw": raw_serialized,
                    "kind": infer_kind(raw_value, value_cell.number_format or formula_cell.number_format)
                }
                if formula:
                    payload["formula"] = formula

                row_cells[str(column_index)] = payload

            if row_cells:
                sheet_rows.append({
                    "index": row_index,
                    "cells": row_cells
                })

        sheets_payload.append({
            "name": sheet_formula.title,
            "path": f"xlsx/{sheet_formula.title}",
            "rows": sheet_rows,
            "nonEmpty": non_empty,
            "formulas": formulas,
            "maxCol": max_col
        })

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": source_path.stem,
        "sheets": sheets_payload
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build workbook_snapshot.json from ЛАЙТ 2.xlsx")
    parser.add_argument("source", help="Source XLSX workbook path")
    parser.add_argument("output", help="Output JSON path")
    args = parser.parse_args()

    source_path = Path(args.source).resolve()
    output_path = Path(args.output).resolve()

    snapshot = build_snapshot(source_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"SNAPSHOT_OK:{output_path}")


if __name__ == "__main__":
    main()
